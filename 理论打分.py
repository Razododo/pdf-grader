# -*- coding: utf-8 -*-
"""
PDF Grader GUI (Preview + Manual scoring + Save CSV + Export Gradebook Excel)

- Read PDFs from ./submissions (or selectable folder)
- Filename: <class_time>-<name>-<student_id>.pdf
  Example: 上午1-2节（830-1000）-蔡俊灏-2023151554.pdf

Features:
- Preview PDF pages (PyMuPDF)
- Manual score input for Q1/Q2 (0-50), optional comment
- Auto total_score = (q1 or 0) + (q2 or 0)
- Save/Update to manual_grades.csv (auto total_score)
- One-click export gradebook_export.xlsx (Excel export)
    * grouped by class_time (separate sheets)
    * sorted by student_id
    * missing submission / missing Q1/Q2 highlighted
- Can choose submissions folder in the UI

Requirements:
  pip install pymupdf pillow pandas openpyxl
"""

import os, re, glob
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk

import fitz  # PyMuPDF

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

import sys
from pathlib import Path
from tkinter import filedialog

def get_app_dir() -> Path:
    """Return the directory where the app is running.
    Works for normal python runs and for PyInstaller-built executables."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

APP_DIR = get_app_dir()

# Default submissions folder (can be changed in UI)
DEFAULT_SUBMISSIONS_DIR = APP_DIR / "submissions"

# Outputs are saved next to the app by default
OUT_CSV = APP_DIR / "manual_grades.csv"
OUT_XLSX = APP_DIR / "gradebook_export.xlsx"

RENDER_DPI = 130  # preview quality; lower for smoother preview on slower machines

# Excel styles
FILL_MISSING = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
FILL_ALERT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # light red
FONT_BOLD = Font(bold=True)
ALIGN_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)
ALIGN_WRAP = Alignment(vertical="top", wrap_text=True)

def parse_filename_meta(filepath: str):
    base = os.path.splitext(os.path.basename(filepath))[0]
    parts = base.split("-")
    meta = {"class_time": "", "name": "", "student_id": ""}

    if len(parts) >= 3:
        meta["student_id"] = parts[-1].strip()
        meta["name"] = parts[-2].strip()
        meta["class_time"] = "-".join(parts[:-2]).strip()
    else:
        m = re.search(r"(\d{8,})", base)
        if m:
            meta["student_id"] = m.group(1)
        meta["class_time"] = base.strip()

    return meta

def clamp_int(s, lo, hi, default=None):
    try:
        v = int(str(s).strip())
        if v < lo: return lo
        if v > hi: return hi
        return v
    except:
        return default

def safe_int_or_zero(x):
    try:
        if x is None: return 0
        s = str(x).strip()
        if s == "" or s.lower() == "nan":
            return 0
        return int(float(s))
    except:
        return 0

def export_gradebook(df: pd.DataFrame, out_xlsx: str):
    """
    Export gradebook grouped by class_time into Excel.
    - One sheet per class_time
    - Sorted by student_id
    - Highlight missing Q1/Q2 (yellow) and missing student_id/name (red)
    """
    # Ensure columns
    cols = ["class_time","name","student_id","q1_score","q2_score","total_score","comment","file"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""

    # Normalize types
    df2 = df.copy()
    df2["student_id_str"] = df2["student_id"].astype(str)
    df2["class_time_str"] = df2["class_time"].astype(str)

    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # If class_time empty, put into "未分组"
    class_times = sorted(set([ct if ct.strip() else "未分组" for ct in df2["class_time_str"].fillna("").tolist()]))

    for ct in class_times:
        sheet_name = ct
        # Excel sheet name limit 31 chars
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        # Avoid duplicates
        base_name = sheet_name
        k = 1
        while sheet_name in wb.sheetnames:
            suffix = f"_{k}"
            sheet_name = (base_name[:31-len(suffix)] + suffix) if len(base_name) + len(suffix) > 31 else base_name + suffix
            k += 1

        ws = wb.create_sheet(title=sheet_name)

        sub = df2[df2["class_time_str"].fillna("").apply(lambda x: x.strip() if isinstance(x,str) else "") == (ct if ct!="未分组" else "")]
        if ct == "未分组":
            sub = df2[df2["class_time_str"].fillna("").apply(lambda x: x.strip() if isinstance(x,str) else "") == ""]

        # sort by student_id (numeric if possible)
        def sort_key_sid(s):
            s = str(s).strip()
            return int(s) if s.isdigit() else 10**18
        sub = sub.copy()
        sub["__sid_sort"] = sub["student_id"].apply(sort_key_sid)
        sub = sub.sort_values(["__sid_sort","student_id","name"], ascending=True).drop(columns="__sid_sort")

        # Keep export columns
        export_cols = ["class_time","name","student_id","q1_score","q2_score","total_score","comment","file"]
        sub_export = sub[export_cols]

        # Write header
        ws.append(export_cols)
        for cell in ws[1]:
            cell.font = FONT_BOLD
            cell.alignment = ALIGN_CENTER

        # Write rows
        for r in dataframe_to_rows(sub_export, index=False, header=False):
            ws.append(r)

        # Formatting: column widths
        col_widths = {
            "A": 24, "B": 10, "C": 14, "D": 10, "E": 10, "F": 12, "G": 40, "H": 30
        }
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # Row styling and highlights
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Align
            for c in row:
                c.alignment = ALIGN_WRAP

            name = str(row[1].value or "").strip()
            sid = str(row[2].value or "").strip()
            q1 = str(row[3].value or "").strip()
            q2 = str(row[4].value or "").strip()

            # Missing critical id/name => red
            if name == "" or sid == "":
                for c in row:
                    c.fill = FILL_ALERT

            # Missing q1/q2 => yellow (not as severe)
            if q1 == "":
                row[3].fill = FILL_MISSING
            if q2 == "":
                row[4].fill = FILL_MISSING

    wb.save(out_xlsx)

class PDFGraderApp:
    def __init__(self, root):
        self.root = root
        root.title("PDF 作业批改器（预览+打分+保存+导出登记表）")

        self.submissions_dir = Path(DEFAULT_SUBMISSIONS_DIR)
        self.pdf_paths = self._scan_pdfs()
        if not self.pdf_paths:
            messagebox.showerror(
                "未找到文件",
                f"在 {self.submissions_dir} 下没有找到 PDF\n\n你可以点击右侧按钮选择 submissions 文件夹。"
            )
            # Quick input state (needed even when no PDFs are loaded yet)
            self._quick_stage = 1
            self._quick_buffer = ""
            # Load grade CSV so export works after choosing folder
            self.df = self._load_existing_csv()
            # Basic state placeholders
            self.idx = 0
            self.doc = None
            self.page_index = 0
            self.current_image_tk = None
            self._build_ui()
            self.var_submissions_dir.set(str(self.submissions_dir))
            self.canvas.delete("all")
            self.page_label.config(text="Page 0/0")
            self._update_status()
            return

        self.df = self._load_existing_csv()

        self.idx = 0
        self.doc = None
        self.page_index = 0
        self.current_image_tk = None

        # Debounce and cache attributes for resize/render
        self._resize_after_id = None
        self._last_canvas_size = (0, 0)
        self._page_render_cache = {}
        self._last_render_key = None
        # Quick score input state: digits -> Enter(Q1), digits -> Enter(Q2), Enter(save+next)
        self._quick_stage = 1  # 1=q1 input, 2=q2 input, 3=ready to save+next
        self._quick_buffer = ""  # digit buffer for current stage
        self._build_ui()
        self._load_pdf(self.idx)

    def _load_existing_csv(self):
        if OUT_CSV.exists():
            try:
                df = pd.read_csv(str(OUT_CSV), encoding="utf-8-sig")
                for col in ["class_time","name","student_id","q1_score","q2_score","total_score","comment","file"]:
                    if col not in df.columns:
                        df[col] = ""
                return df
            except Exception:
                pass
        return pd.DataFrame(columns=["class_time","name","student_id","q1_score","q2_score","total_score","comment","file"])

    def _scan_pdfs(self):
        if not self.submissions_dir.exists():
            return []
        return sorted([str(p) for p in self.submissions_dir.glob("*.pdf")])

    def _build_ui(self):
        self.root.geometry("1280x820")

        self.main = ttk.Frame(self.root, padding=8)
        self.main.pack(fill="both", expand=True)

        # Draggable split panes (preview on the left, scoring on the right)
        self.paned = ttk.PanedWindow(self.main, orient="horizontal")
        self.paned.pack(fill="both", expand=True)

        self.left = ttk.Frame(self.paned)
        self.right = ttk.Frame(self.paned, width=380)

        # Add panes; set initial weights so the left takes most space
        self.paned.add(self.left, weight=3)
        self.paned.add(self.right, weight=1)

        # Set an initial sash position (approx 70% left / 30% right)
        self.root.after(0, lambda: self._set_initial_sash())

        # Preview
        self.canvas = tk.Canvas(self.left, bg="black")
        self.canvas.pack(fill="both", expand=True)

        # Page controls
        self.page_bar = ttk.Frame(self.left)
        self.page_bar.pack(fill="x", pady=(8,0))

        ttk.Button(self.page_bar, text="上一页", command=self.prev_page).pack(side="left")
        self.page_label = ttk.Label(self.page_bar, text="Page 1/1", width=20)
        self.page_label.pack(side="left", padx=8)
        ttk.Button(self.page_bar, text="下一页", command=self.next_page).pack(side="left")

        # Meta
        meta_frame = ttk.LabelFrame(self.right, text="学生信息", padding=10)
        meta_frame.pack(fill="x", pady=(0,10))

        self.var_class_time = tk.StringVar()
        self.var_name = tk.StringVar()
        self.var_student_id = tk.StringVar()
        self.var_file = tk.StringVar()
        self.var_submissions_dir = tk.StringVar(value=str(self.submissions_dir))

        ttk.Label(meta_frame, text="上课时间：").grid(row=0, column=0, sticky="w")
        ttk.Label(meta_frame, textvariable=self.var_class_time, wraplength=330).grid(row=0, column=1, sticky="w")

        ttk.Label(meta_frame, text="姓名：").grid(row=1, column=0, sticky="w")
        ttk.Label(meta_frame, textvariable=self.var_name).grid(row=1, column=1, sticky="w")

        ttk.Label(meta_frame, text="学号：").grid(row=2, column=0, sticky="w")
        ttk.Label(meta_frame, textvariable=self.var_student_id).grid(row=2, column=1, sticky="w")

        ttk.Label(meta_frame, text="文件：").grid(row=3, column=0, sticky="w")
        ttk.Label(meta_frame, textvariable=self.var_file, wraplength=330).grid(row=3, column=1, sticky="w")

        ttk.Label(meta_frame, text="作业文件夹：").grid(row=4, column=0, sticky="w")
        ttk.Label(meta_frame, textvariable=self.var_submissions_dir, wraplength=330).grid(row=4, column=1, sticky="w")
        ttk.Button(meta_frame, text="选择…", command=self.choose_submissions_dir).grid(row=5, column=1, sticky="w", pady=(6,0))

        # Scoring
        score_frame = ttk.LabelFrame(self.right, text="打分（0-50）", padding=10)
        score_frame.pack(fill="x", pady=(0,10))

        self.entry_q1 = ttk.Entry(score_frame, width=10)
        self.entry_q2 = ttk.Entry(score_frame, width=10)
        self.var_total = tk.StringVar(value="0")

        ttk.Label(score_frame, text="题1分数：").grid(row=0, column=0, sticky="w")
        self.entry_q1.grid(row=0, column=1, sticky="w")

        ttk.Label(score_frame, text="题2分数：").grid(row=1, column=0, sticky="w")
        self.entry_q2.grid(row=1, column=1, sticky="w")

        ttk.Label(score_frame, text="总分（自动）：").grid(row=2, column=0, sticky="w", pady=(6,0))
        ttk.Label(score_frame, textvariable=self.var_total, font=("Arial", 14, "bold")).grid(row=2, column=1, sticky="w", pady=(6,0))

        ttk.Label(score_frame, text="提示：只写其中一题，另一题可留空（按0计）。",
                  wraplength=330).grid(row=3, column=0, columnspan=2, sticky="w", pady=(8,0))

        # Live total update
        def on_score_change(*_):
            q1 = clamp_int(self.entry_q1.get(), 0, 50, default=None)
            q2 = clamp_int(self.entry_q2.get(), 0, 50, default=None)
            total = (q1 if q1 is not None else 0) + (q2 if q2 is not None else 0)
            self.var_total.set(str(total))
        self.entry_q1.bind("<KeyRelease>", lambda e: on_score_change())
        self.entry_q2.bind("<KeyRelease>", lambda e: on_score_change())

        # Comment
        comment_frame = ttk.LabelFrame(self.right, text="评语（可选）", padding=10)
        comment_frame.pack(fill="both", expand=True, pady=(0,10))

        self.text_comment = tk.Text(comment_frame, height=10, wrap="word")
        self.text_comment.pack(fill="both", expand=True)

        # Action buttons
        action_frame = ttk.Frame(self.right)
        action_frame.pack(fill="x")

        ttk.Button(action_frame, text="上一份", command=self.prev_pdf).pack(side="left")
        ttk.Button(action_frame, text="保存(Ctrl+S)", command=self.save_current).pack(side="left", padx=6)
        ttk.Button(action_frame, text="保存并下一份", command=self.save_and_next).pack(side="left", padx=6)
        ttk.Button(action_frame, text="下一份", command=self.next_pdf).pack(side="left", padx=6)

        export_frame = ttk.Frame(self.right)
        export_frame.pack(fill="x", pady=(8,0))
        ttk.Button(export_frame, text="一键导出登记表（Excel）", command=self.export_excel).pack(side="left")

        self.status = ttk.Label(self.right, text="", foreground="gray")
        self.status.pack(fill="x", pady=(8,0))

        # Shortcuts
        self.root.bind("<Left>", lambda e: self.prev_page())
        self.root.bind("<Right>", lambda e: self.next_page())
        self.root.bind("<Up>", lambda e: self.prev_pdf())
        self.root.bind("<Down>", lambda e: self.next_pdf())
        self.root.bind("<Control-s>", lambda e: self.save_current())
        # Quick numeric input bindings (ignored when typing in comment box)
        self.root.bind("<Key>", self._on_keypress)
        self.root.bind("<Return>", self._on_enter)
        self.root.bind("<KP_Enter>", self._on_enter)
        self.root.bind("<BackSpace>", self._on_backspace)
        self.root.bind("<Escape>", self._on_escape)
        
    def _load_pdf(self, idx):
        if not self.pdf_paths:
            return
        if self.doc is not None:
            self.doc.close()

        self.idx = max(0, min(idx, len(self.pdf_paths)-1))
        path = self.pdf_paths[self.idx]

        self.doc = fitz.open(path)
        self.page_index = 0

        meta = parse_filename_meta(path)
        self.var_class_time.set(meta["class_time"])
        self.var_name.set(meta["name"])
        self.var_student_id.set(meta["student_id"])
        self.var_file.set(os.path.basename(path))

        self._populate_existing(meta["student_id"])
        self._render_page()
        self._update_status()
        self._reset_quick(stage=1)

    def _populate_existing(self, student_id):
        self.entry_q1.delete(0, "end")
        self.entry_q2.delete(0, "end")
        self.text_comment.delete("1.0", "end")
        self.var_total.set("0")

        sid = str(student_id).strip()
        if not sid:
            return

        hit = self.df[self.df["student_id"].astype(str) == sid]
        if len(hit) > 0:
            row = hit.iloc[-1]
            q1v = "" if pd.isna(row.get("q1_score", "")) else str(row.get("q1_score", "")).strip()
            q2v = "" if pd.isna(row.get("q2_score", "")) else str(row.get("q2_score", "")).strip()
            if q1v != "": self.entry_q1.insert(0, q1v)
            if q2v != "": self.entry_q2.insert(0, q2v)

            total = safe_int_or_zero(q1v) + safe_int_or_zero(q2v)
            self.var_total.set(str(total))

            c = "" if pd.isna(row.get("comment", "")) else str(row.get("comment", ""))
            self.text_comment.insert("1.0", c)

    def _render_page(self):
        if self.doc is None:
            return

        n_pages = len(self.doc)
        self.page_index = max(0, min(self.page_index, n_pages-1))

        canvas_w = max(100, self.canvas.winfo_width())
        canvas_h = max(100, self.canvas.winfo_height())
        cache_key = (self.var_file.get(), self.page_index, RENDER_DPI)
        if self._last_render_key == cache_key and self._last_canvas_size == (canvas_w, canvas_h):
            return
        self._last_render_key = cache_key
        self._last_canvas_size = (canvas_w, canvas_h)

        # Caching: render PDF page at DPI once, then resize for display
        if cache_key in self._page_render_cache:
            img = self._page_render_cache[cache_key]
        else:
            page = self.doc.load_page(self.page_index)
            zoom = RENDER_DPI / 72.0
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            self._page_render_cache[cache_key] = img

        img_ratio = img.width / img.height
        canvas_ratio = canvas_w / canvas_h

        if img_ratio > canvas_ratio:
            new_w = canvas_w
            new_h = int(canvas_w / img_ratio)
        else:
            new_h = canvas_h
            new_w = int(canvas_h * img_ratio)

        img_resized = img.resize((new_w, new_h), Image.LANCZOS)
        self.current_image_tk = ImageTk.PhotoImage(img_resized)

        self.canvas.delete("all")
        self.canvas.create_image(canvas_w//2, canvas_h//2, image=self.current_image_tk, anchor="center")
        self.page_label.config(text=f"Page {self.page_index+1}/{n_pages}")

    def _update_status(self):
        self._refresh_quick_hint()

    def _set_initial_sash(self):
        try:
            self.root.update_idletasks()
            w = self.root.winfo_width()
            if hasattr(self, "paned") and w > 200:
                self.paned.sashpos(0, int(w * 0.72))
        except Exception:
            pass

    def _focus_in_comment(self) -> bool:
        try:
            w = self.root.focus_get()
            return w is self.text_comment
        except Exception:
            return False

    def _quick_stage_label(self) -> str:
        if self._quick_stage == 1:
            return "题1"
        if self._quick_stage == 2:
            return "题2"
        return "保存"

    def _refresh_quick_hint(self):
        base = (
            f"进度：{(self.idx+1) if self.pdf_paths else 0}/{len(self.pdf_paths)}   "
            f"文件夹：{self.submissions_dir}   快捷键：←→翻页  ↑↓切作业  Ctrl+S保存"
        )
        hint = (
            f"   快捷输入：当前{self._quick_stage_label()} 输入[{self._quick_buffer}]"
            f"（数字输入，回车确认；再回车保存下一份；Backspace退格；Esc重置）"
        )
        if hasattr(self, "status"):
            self.status.config(text=base + hint)

    def _reset_quick(self, stage: int = 1):
        self._quick_stage = stage
        self._quick_buffer = ""
        self._refresh_quick_hint()

    def _preview_quick_value(self, stage: int):
        """Update entry preview based on current buffer (or clear if empty)."""
        if stage not in (1, 2):
            return

        if self._quick_buffer.strip() == "":
            if stage == 1:
                self.entry_q1.delete(0, "end")
            else:
                self.entry_q2.delete(0, "end")
            self._compute_total()
            self._refresh_quick_hint()
            return

        try:
            preview_val = int(self._quick_buffer)
        except Exception:
            preview_val = 0
        preview_val = max(0, min(50, preview_val))

        if stage == 1:
            self.entry_q1.delete(0, "end")
            self.entry_q1.insert(0, str(preview_val))
        else:
            self.entry_q2.delete(0, "end")
            self.entry_q2.insert(0, str(preview_val))

        self._compute_total()
        self._refresh_quick_hint()

    def _commit_quick_buffer(self):
        """Commit current buffer into Q1 or Q2 depending on stage; empty buffer commits 0."""
        if self._quick_stage not in (1, 2):
            return

        if self._quick_buffer.strip() == "":
            val = 0
        else:
            try:
                val = int(self._quick_buffer)
            except Exception:
                val = 0

        val = max(0, min(50, val))

        if self._quick_stage == 1:
            self.entry_q1.delete(0, "end")
            self.entry_q1.insert(0, str(val))
        else:
            self.entry_q2.delete(0, "end")
            self.entry_q2.insert(0, str(val))

        self._compute_total()

    def _on_keypress(self, event):
        if self._focus_in_comment():
            return
        if self._quick_stage not in (1, 2):
            return

        ch = event.char
        if not ch or not ch.isdigit():
            return

        # Buffer supports 1-2 digits; if already 2 digits, restart with new digit
        if len(self._quick_buffer) >= 2:
            self._quick_buffer = ch
        else:
            self._quick_buffer += ch

        self._preview_quick_value(self._quick_stage)

    def _on_backspace(self, event=None):
        if self._focus_in_comment():
            return
        if self._quick_stage not in (1, 2):
            return
        if self._quick_buffer:
            self._quick_buffer = self._quick_buffer[:-1]
        self._preview_quick_value(self._quick_stage)

    def _on_enter(self, event=None):
        if self._focus_in_comment():
            return

        if self._quick_stage == 1:
            self._commit_quick_buffer()  # empty => 0
            self._quick_stage = 2
            self._quick_buffer = ""
            self._refresh_quick_hint()
            return

        if self._quick_stage == 2:
            self._commit_quick_buffer()  # empty => 0
            self._quick_stage = 3
            self._quick_buffer = ""
            self._refresh_quick_hint()
            return

        # Stage 3
        self.save_and_next()
        self._reset_quick(stage=1)

    def _on_escape(self, event=None):
        if self._focus_in_comment():
            return
        self._reset_quick(stage=1)

    def prev_page(self):
        if self.doc and self.page_index > 0:
            self.page_index -= 1
            self._render_page()

    def next_page(self):
        if self.doc and self.page_index < len(self.doc)-1:
            self.page_index += 1
            self._render_page()

    def prev_pdf(self):
        if self.idx > 0:
            self._load_pdf(self.idx - 1)

    def next_pdf(self):
        if self.idx < len(self.pdf_paths)-1:
            self._load_pdf(self.idx + 1)

    def _compute_total(self):
        q1 = clamp_int(self.entry_q1.get(), 0, 50, default=None)
        q2 = clamp_int(self.entry_q2.get(), 0, 50, default=None)
        total = (q1 if q1 is not None else 0) + (q2 if q2 is not None else 0)
        self.var_total.set(str(total))
        return q1, q2, total

    def save_current(self):
        q1, q2, total = self._compute_total()

        if q1 is None and q2 is None:
            messagebox.showwarning("未填写分数", "题1/题2至少填写一个分数（0-50）。")
            return False

        record = {
            "class_time": self.var_class_time.get(),
            "name": self.var_name.get(),
            "student_id": str(self.var_student_id.get()).strip(),
            "q1_score": "" if q1 is None else q1,
            "q2_score": "" if q2 is None else q2,
            "total_score": total,
            "comment": self.text_comment.get("1.0", "end").strip(),
            "file": self.var_file.get(),
        }

        sid = record["student_id"]
        if sid:
            mask = self.df["student_id"].astype(str) == sid
            if mask.any():
                last_idx = self.df[mask].index[-1]
                for k, v in record.items():
                    self.df.at[last_idx, k] = v
            else:
                self.df = pd.concat([self.df, pd.DataFrame([record])], ignore_index=True)
        else:
            self.df = pd.concat([self.df, pd.DataFrame([record])], ignore_index=True)

        self.df.to_csv(str(OUT_CSV), index=False, encoding="utf-8-sig")
        self.status.config(text=f"已保存：{str(OUT_CSV)}（{record['name']} {record['student_id']} 总分={total}）")
        return True

    def save_and_next(self):
        ok = self.save_current()
        if not ok:
            return
        if self.idx < len(self.pdf_paths)-1:
            self._load_pdf(self.idx + 1)
        else:
            messagebox.showinfo("完成", f"已经是最后一份。成绩已保存到 {str(OUT_CSV)}")

    def export_excel(self):
        # Always save current (optional)
        self.save_current()

        try:
            export_gradebook(self.df, str(OUT_XLSX))
            messagebox.showinfo("导出成功", f"已导出：{str(OUT_XLSX)}\n（按上课时间分sheet、按学号排序、缺题标色）")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出Excel失败：\n{e}")

    def choose_submissions_dir(self):
        selected = filedialog.askdirectory(title="选择 submissions 文件夹")
        if not selected:
            return
        self.submissions_dir = Path(selected)
        self.var_submissions_dir.set(str(self.submissions_dir))
        self.pdf_paths = self._scan_pdfs()
        if not self.pdf_paths:
            messagebox.showwarning("未找到PDF", f"在 {self.submissions_dir} 下没有找到 PDF")
            self.idx = 0
            if self.doc is not None:
                self.doc.close()
                self.doc = None
            self.canvas.delete("all")
            self.page_label.config(text="Page 0/0")
            self.var_file.set("")
            self.var_name.set("")
            self.var_student_id.set("")
            self.var_class_time.set("")
            self._update_status()
            return
        # Load the first PDF in the new folder
        self._load_pdf(0)

def main():
    root = tk.Tk()
    app = PDFGraderApp(root)

    # redraw on resize (debounced)
    def on_resize(_event):
        if getattr(app, "_resize_after_id", None) is not None:
            try:
                root.after_cancel(app._resize_after_id)
            except Exception:
                pass
        app._resize_after_id = root.after(180, app._render_page)

    root.bind("<Configure>", on_resize)
    root.mainloop()

if __name__ == "__main__":
    main()